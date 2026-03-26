#!/usr/bin/env python3
"""
Lê NFSe ABRASF (1.00 e 2.02; tolera namespace incorreto nos XMLs)
e gera relatório XLSX com emissão, cancelamento, tomador, valor bruto e competência.
"""

from __future__ import annotations

import argparse
import io
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

COLUNAS = [
    "arquivo",
    "numero_nfse",
    "data_emissao",
    "data_cancelamento",
    "nome_cliente",
    "valor_bruto",
    "competencia",
    "erro",
]

# Formatos de exibição no Excel (pt-BR)
FMT_DATA = "dd/mm/yyyy"
FMT_VALOR = "#.##0,00"


def _texto(el: ET.Element | None) -> str:
    if el is None or el.text is None:
        return ""
    return el.text.strip()


def _parse_iso_para_datetime(s: str) -> datetime | None:
    if not s or not s.strip():
        return None
    s = s.strip().replace("Z", "")
    if "+" in s:
        s = s.split("+", 1)[0]
    if "T" not in s:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            return None
    idx_t = s.index("T")
    idx_dot = s.rfind(".")
    if idx_dot > idx_t:
        prefix = s[:idx_dot]
        frac = "".join(c for c in s[idx_dot + 1 :] if c.isdigit())[:6]
        frac = (frac + "000000")[:6]
        s_fmt = f"{prefix}.{frac}"
        try:
            return datetime.strptime(s_fmt, "%Y-%m-%dT%H:%M:%S.%f")
        except ValueError:
            pass
    try:
        return datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S")
    except ValueError:
        return None


def _parse_iso_para_date(s: str) -> date | None:
    dt = _parse_iso_para_datetime(s)
    if dt is None:
        return None
    return dt.date()


def _parse_valor(s: str) -> float | None:
    if not s:
        return None
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


def _tag_local(tag: str) -> str:
    if tag.startswith("{"):
        return tag[tag.index("}") + 1 :]
    return tag


def _child(parent: ET.Element, local: str) -> ET.Element | None:
    for ch in parent:
        if _tag_local(ch.tag) == local:
            return ch
    return None


def _find_first_descendant(parent: ET.Element, local: str) -> ET.Element | None:
    for el in parent.iter():
        if _tag_local(el.tag) == local:
            return el
    return None


def _text_child(parent: ET.Element | None, local: str) -> str:
    if parent is None:
        return ""
    return _texto(_child(parent, local))


def _inf_declaracao(inf: ET.Element) -> ET.Element | None:
    dps = _child(inf, "DeclaracaoPrestacaoServico")
    if dps is None:
        return None
    return _child(dps, "InfDeclaracaoPrestacaoServico")


def _valor_bruto_de_servico(serv: ET.Element | None) -> float | None:
    if serv is None:
        return None
    vals = _child(serv, "Valores")
    if vals is None:
        return None
    return _parse_valor(_texto(_child(vals, "ValorServicos")))


def _data_cancelamento_root(root: ET.Element) -> datetime | None:
    bloco = _child(root, "NfseCancelamento")
    if bloco is None:
        return None
    conf = _child(bloco, "Confirmacao")
    if conf is None:
        return None
    dh = _child(conf, "DataHora")
    if dh is None:
        dh = _child(conf, "DataHoraCancelamento")
    return _parse_iso_para_datetime(_texto(dh))


def extrair_de_tree(tree: ET.ElementTree, nome_arquivo: str) -> dict[str, Any]:
    linha: dict[str, Any] = {c: "" for c in COLUNAS}
    linha["arquivo"] = nome_arquivo
    root = tree.getroot()
    inf = _find_first_descendant(root, "InfNfse")
    if inf is None:
        linha["erro"] = "InfNfse não encontrado (não parece NFSe ABRASF)."
        return linha

    linha["numero_nfse"] = _text_child(inf, "Numero")
    linha["data_emissao"] = _parse_iso_para_datetime(_text_child(inf, "DataEmissao"))

    dps = _inf_declaracao(inf)
    if dps is not None:
        linha["competencia"] = _parse_iso_para_date(_text_child(dps, "Competencia"))
        tom = _child(dps, "Tomador")
        if tom is not None:
            linha["nome_cliente"] = _text_child(tom, "RazaoSocial")
        linha["valor_bruto"] = _valor_bruto_de_servico(_child(dps, "Servico"))
    else:
        linha["competencia"] = _parse_iso_para_date(_text_child(inf, "Competencia"))
        ts = _child(inf, "TomadorServico")
        if ts is not None:
            linha["nome_cliente"] = _text_child(ts, "RazaoSocial")
        linha["valor_bruto"] = _valor_bruto_de_servico(_child(inf, "Servico"))

    linha["data_cancelamento"] = _data_cancelamento_root(root)

    return linha


def extrair_linha(path: Path) -> dict[str, Any]:
    linha: dict[str, Any] = {c: "" for c in COLUNAS}
    linha["arquivo"] = path.name
    try:
        tree = ET.parse(path)
    except ET.ParseError as e:
        linha["erro"] = f"XML inválido: {e}"
        return linha
    return extrair_de_tree(tree, path.name)


def extrair_linha_de_bytes(nome: str, data: bytes) -> dict[str, Any]:
    linha: dict[str, Any] = {c: "" for c in COLUNAS}
    linha["arquivo"] = nome
    try:
        tree = ET.parse(io.BytesIO(data))
    except ET.ParseError as e:
        linha["erro"] = f"XML inválido: {e}"
        return linha
    return extrair_de_tree(tree, nome)


TITULOS_PT = {
    "arquivo": "Arquivo",
    "numero_nfse": "Nº NFSe",
    "data_emissao": "Data emissão",
    "data_cancelamento": "Data cancelamento",
    "nome_cliente": "Nome cliente",
    "valor_bruto": "Valor bruto",
    "competencia": "Competência",
    "erro": "Erro",
}


def _montar_workbook(linhas: list[dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "NFSe"

    for col, key in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col, value=TITULOS_PT[key])
        cell.font = Font(bold=True)

    for row_idx, linha in enumerate(linhas, start=2):
        for col, key in enumerate(COLUNAS, start=1):
            val = linha.get(key, "")
            c = ws.cell(row=row_idx, column=col)

            if key in ("data_emissao", "data_cancelamento", "competencia"):
                if isinstance(val, datetime):
                    c.value = val
                    c.number_format = FMT_DATA
                elif isinstance(val, date):
                    c.value = val
                    c.number_format = FMT_DATA
                else:
                    c.value = val if val else None
            elif key == "valor_bruto":
                if isinstance(val, (int, float)):
                    c.value = float(val)
                    c.number_format = FMT_VALOR
                else:
                    c.value = None
            else:
                c.value = val if val != "" else None

            if key == "nome_cliente":
                c.alignment = Alignment(wrap_text=True, vertical="top")

    for col in range(1, len(COLUNAS) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(
            48, max(12, len(TITULOS_PT[COLUNAS[col - 1]]) + 2)
        )

    return wb


def xlsx_em_bytes(linhas: list[dict[str, Any]]) -> bytes:
    bio = io.BytesIO()
    _montar_workbook(linhas).save(bio)
    return bio.getvalue()


def _escrever_xlsx(linhas: list[dict[str, Any]], caminho: Path) -> None:
    caminho.parent.mkdir(parents=True, exist_ok=True)
    _montar_workbook(linhas).save(caminho)


def gerar_relatorio(pasta_xml: Path, saida_xlsx: Path) -> list[dict[str, Any]]:
    if not pasta_xml.is_dir():
        raise SystemExit(f"Pasta não encontrada: {pasta_xml}")

    arquivos = sorted(pasta_xml.glob("*.xml"))
    if not arquivos:
        raise SystemExit(f"Nenhum .xml em {pasta_xml}")

    linhas = [extrair_linha(p) for p in arquivos]
    _escrever_xlsx(linhas, saida_xlsx)
    return linhas


def main() -> None:
    base = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Relatório NFSe ABRASF a partir de XMLs.")
    parser.add_argument(
        "--pasta",
        type=Path,
        default=base / "XMLs",
        help="Pasta com arquivos .xml (padrão: ./XMLs ao lado deste script)",
    )
    parser.add_argument(
        "--saida",
        type=Path,
        default=base / "relatorio_nfse.xlsx",
        help="Arquivo XLSX de saída",
    )
    args = parser.parse_args()
    linhas = gerar_relatorio(args.pasta.resolve(), args.saida.resolve())
    erros = sum(1 for r in linhas if r.get("erro"))
    print(f"Processados: {len(linhas)} arquivo(s). Planilha: {args.saida.resolve()}")
    if erros:
        print(f"Aviso: {erros} registro(s) com erro — veja coluna 'Erro' na planilha.")


if __name__ == "__main__":
    main()
